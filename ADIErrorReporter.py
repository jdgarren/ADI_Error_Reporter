import subprocess
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, TimeoutException
error = (NoSuchElementException, ElementNotInteractableException, TimeoutException)
##from openpyxl import worksheet
wb1 = load_workbook(r'C:\Users\garren-james\Python Projects\ADINotExist.xlsx')
ws1 = wb1['Sheet1']
loc = (r'C:\Users\garren-james\ProdverLogs')
user = loc + "\\ADICred.txt"
signin = open(user, 'r')
user = signin.read().splitlines()
rn = input('What row to start with? ')
en = input('What number to stop at? ')
rn1 = int(rn)
en1 = int(en)
while rn1 <= en1:
    browser = webdriver.Firefox(executable_path=r'C:\Users\garren-james\AppData\Local\geckodriver')
    browser.get('http://imaging.dcf.state.fl.us/Default.aspx')
    loc1 = ws1.cell(row=rn1, column=1).value
    htmlElem = browser.find_element_by_id('ctl07_tbUsername')
    htmlElem.send_keys(user[0])
    htmlElem = browser.find_element_by_id('ctl07_tbPassword')
    htmlElem.send_keys(user[1])
    htmlElem = browser.find_element_by_id('ctl07_lbSubmit')
    htmlElem.click()
    htmlElem = browser.find_element_by_id('ucMenu_lblPsnDocSearch')
    htmlElem.click()
    htmlElem = browser.find_element_by_id('ctl07_ITDocNumber')
    htmlElem.send_keys(loc1)
    htmlElem = browser.find_element_by_id('ctl07_BTNSearch')
    htmlElem.click()
    htmlElem = browser.find_element_by_id('ctl07_grdView_lbReportError_0')
    htmlElem.click()
    select = Select(browser.find_element_by_id('ctl07_ucSendWorkGroup_ddlRegion'))
    select.select_by_value('99')
    select = Select(browser.find_element_by_id('ctl07_ucSendWorkGroup_ddlWorkGroupType'))
    select.select_by_value('74')
    select = Select(browser.find_element_by_id('ctl07_ucSendWorkGroup_ddlWorkGroups'))
    select.select_by_value('2442')
    htmlElem = browser.find_element_by_id('ctl07_ucSendWorkGroup_btnSend')
    htmlElem.click()
    browser.quit()
    rn1 = rn1+1